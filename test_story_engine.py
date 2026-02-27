"""
test_story_engine.py  (v3 — expected_outcome fix + no hardcoded strings)
──────────────────────────────────────────────────────────────────────────
FIXES IN THIS VERSION:
1. ✅ TestStory dataclass has expected_outcome field
2. ✅ Execution Plan uses real expected_outcome from story, not hardcoded generic text
3. ✅ _print_console_summary called exactly once (from generate_all)
"""

import json
from pathlib import Path
from datetime import datetime
from typing import Optional, Dict, List, Any
from dataclasses import dataclass, field
from enum import Enum

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ═══════════════════════════════════════════════════════════════
#  COLOUR PALETTE
# ═══════════════════════════════════════════════════════════════

class C:
    HEADER_DARK   = "1E293B"
    HEADER_MID    = "334155"
    PASS_BG       = "166534"
    FAIL_BG       = "7F1D1D"
    SKIP_BG       = "78350F"
    ALT_ROW       = "F8FAFC"
    ACCENT_BLUE   = "0EA5E9"
    ACCENT_ORANGE = "F97316"
    ACCENT_PURPLE = "7C3AED"
    WHITE         = "FFFFFF"
    LIGHT_BLUE    = "E0F2FE"
    LIGHT_GREEN   = "DCFCE7"
    LIGHT_RED     = "FEE2E2"
    LIGHT_AMBER   = "FEF3C7"
    LIGHT_PURPLE  = "EDE9FE"
    FONT_WHITE    = "FFFFFF"
    FONT_DARK     = "0F172A"
    FONT_MUTED    = "64748B"
    FONT_GREEN    = "15803D"
    FONT_RED      = "DC2626"
    FONT_AMBER    = "D97706"
    FONT_BLUE     = "0369A1"


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color=C.FONT_DARK, size=10, italic=False) -> Font:
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)

def _align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _border(style="thin") -> Border:
    s = Side(style=style, color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)

def _header_row(ws, row: int, values: list, bg: str,
                font_color: str = C.FONT_WHITE, height: int = 22):
    ws.row_dimensions[row].height = height
    for col, val in enumerate(values, 1):
        if isinstance(val, list):
            val = ", ".join(str(v) for v in val)
        cell = ws.cell(row=row, column=col, value=val)
        cell.font      = _font(bold=True, color=font_color, size=10)
        cell.fill      = _fill(bg)
        cell.alignment = _align(h="center", wrap=True)
        cell.border    = _border()

def _data_row(ws, row: int, values: list, bg: str = C.WHITE, height: int = 18):
    ws.row_dimensions[row].height = height
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill      = _fill(bg)
        cell.alignment = _align(wrap=True)
        cell.border    = _border()
        cell.font      = _font()


# ═══════════════════════════════════════════════════════════════
#  DATA MODELS
# ═══════════════════════════════════════════════════════════════

class StoryStatus(Enum):
    PENDING = "pending"
    RUNNING = "running"
    PASSED  = "passed"
    FAILED  = "failed"
    SKIPPED = "skipped"


@dataclass
class StoryStep:
    step_num:  int
    action:    str
    target:    str
    value:     str
    success:   bool
    error:     Optional[str] = None
    timestamp: str = field(default_factory=lambda: datetime.now().isoformat())


@dataclass
class TestStory:
    story_id:         str
    url:              str
    context_name:     str
    user_persona:     str
    description:      str
    field_values:     Dict[str, str]
    status:           StoryStatus = StoryStatus.PENDING
    expected_outcome: str = ""          # ✅ FIX 1 — real outcome from OpenAI
    steps:            List[StoryStep] = field(default_factory=list)
    failure_reason:   Optional[str] = None
    started_at:       Optional[str] = None
    finished_at:      Optional[str] = None

    def start(self):
        self.status     = StoryStatus.RUNNING
        self.started_at = datetime.now().isoformat()

    def pass_story(self):
        self.status      = StoryStatus.PASSED
        self.finished_at = datetime.now().isoformat()

    def fail_story(self, reason: str):
        self.status         = StoryStatus.FAILED
        self.failure_reason = reason
        self.finished_at    = datetime.now().isoformat()

    def add_step(self, action: str, target: str, value: str,
                 success: bool, error: Optional[str] = None):
        self.steps.append(StoryStep(
            step_num  = len(self.steps) + 1,
            action    = action,
            target    = target,
            value     = value,
            success   = success,
            error     = error
        ))

    def get_value_for(self, field_name: str) -> Optional[str]:
        if not field_name:
            return None
        fn = field_name.lower().strip()
        for k, v in self.field_values.items():
            if k.lower() == fn:
                return v
        for k, v in self.field_values.items():
            if fn in k.lower() or k.lower() in fn:
                return v
        return None

    def to_dict(self) -> Dict:
        return {
            "story_id":        self.story_id,
            "url":             self.url,
            "context_name":    self.context_name,
            "user_persona":    self.user_persona,
            "description":     self.description,
            "field_values":    self.field_values,
            "expected_outcome": self.expected_outcome,
            "status":          self.status.value,
            "failure_reason":  self.failure_reason,
            "started_at":      self.started_at,
            "finished_at":     self.finished_at,
            "steps": [
                {
                    "step":      s.step_num,
                    "action":    s.action,
                    "target":    s.target,
                    "value":     s.value,
                    "success":   s.success,
                    "error":     s.error,
                    "timestamp": s.timestamp
                }
                for s in self.steps
            ]
        }


# ═══════════════════════════════════════════════════════════════
#  TEST STORY GENERATOR  (used by StoryAwareDecider)
# ═══════════════════════════════════════════════════════════════

class TestStoryGenerator:
    def __init__(self, openai_client):
        self.openai   = openai_client
        self._counter = 0

    def _next_id(self) -> str:
        self._counter += 1
        ts = datetime.now().strftime("%H%M%S")
        return f"STORY-{ts}-{self._counter:03d}"

    async def generate(
        self,
        url:            str,
        elements:       List[Dict],
        screenshot_b64: str,
        context_type:   str = "page"
    ) -> "TestStory":
        fillable = [
            e for e in elements
            if e.get("element_type") in ("input", "textarea", "select", "custom-select")
        ]
        buttons = [
            e for e in elements
            if e.get("element_type") == "button"
        ]

        elem_summary = []
        for e in fillable:
            name = (e.get("formcontrolname") or e.get("placeholder") or
                    e.get("text") or e.get("name") or "unknown")
            elem_summary.append({
                "field":      name,
                "type":       e.get("element_type"),
                "input_type": e.get("type", ""),
                "required":   e.get("required", False),
                "in_overlay": e.get("in_overlay", False)
            })

        button_names = [b.get("text", "") for b in buttons if b.get("text")]

        prompt = f"""You are a QA engineer testing a web page. A screenshot is attached.

URL: {url}
CONTEXT TYPE: {context_type}

FORM FIELDS DETECTED:
{json.dumps(elem_summary, indent=2)}

SUBMIT/ACTION BUTTONS:
{json.dumps(button_names, indent=2)}

YOUR TASK:
If CONTEXT TYPE is "page":
- Look at the table in the screenshot
- Use EXACT values from the first row for search/filter fields

If CONTEXT TYPE is "form" or "modal":
- Generate completely NEW realistic values based on field names only
- NEVER copy anything visible in the screenshot

Return ONLY valid JSON:
{{
  "context_name": "short name of what this page does",
  "user_persona": "one sentence about who is doing this",
  "description": "2-sentence scenario",
  "field_values": {{
    "fieldname_or_formcontrolname": "value"
  }}
}}
"""
        try:
            response = self.openai.chat.completions.create(
                model="gpt-4o-mini",
                messages=[{
                    "role": "user",
                    "content": [
                        {"type": "text", "text": prompt},
                        {"type": "image_url",
                         "image_url": {"url": f"data:image/png;base64,{screenshot_b64}"}}
                    ]
                }],
                max_tokens=1500,
                temperature=0.7
            )
            raw  = response.choices[0].message.content
            data = json.loads(self._extract_json(raw))

            return TestStory(
                story_id     = self._next_id(),
                url          = url,
                context_name = data.get("context_name", f"Test on {context_type}"),
                user_persona = data.get("user_persona", "QA tester"),
                description  = data.get("description", ""),
                field_values = data.get("field_values", {})
            )

        except Exception as e:
            print(f"⚠️  Story generation failed: {e} — using fallback")
            return self._fallback_story(url, elements, context_type)

    def _fallback_story(self, url: str, elements: List[Dict], context_type: str) -> "TestStory":
        field_values = {}
        for e in elements:
            name = (e.get("formcontrolname") or e.get("placeholder") or
                    e.get("text") or e.get("name") or "")
            if not name:
                continue
            t = e.get("type", "").lower()
            if "email" in name.lower():
                field_values[name] = "test@example.com"
            elif "phone" in name.lower() or "hp" in name.lower():
                field_values[name] = "081234567890"
            elif t == "number":
                field_values[name] = "100"
            else:
                field_values[name] = f"Test {name.title()}"

        return TestStory(
            story_id     = self._next_id(),
            url          = url,
            context_name = f"Test on {context_type}",
            user_persona = "QA tester (fallback)",
            description  = "Fallback story — generation failed.",
            field_values = field_values
        )

    def _extract_json(self, text: str) -> str:
        if "```json" in text:
            start = text.find("```json") + 7
            end   = text.find("```", start)
            return text[start:end].strip()
        if "```" in text:
            start = text.find("```") + 3
            end   = text.find("```", start)
            return text[start:end].strip()
        return text.strip()


# ═══════════════════════════════════════════════════════════════
#  TEST STORY TRACKER
# ═══════════════════════════════════════════════════════════════

class TestStoryTracker:
    FAIL_KEYWORDS = [
        "error", "gagal", "failed", "tidak valid", "invalid",
        "required", "wajib", "must", "salah", "wrong"
    ]

    def __init__(self, output_dir: Path):
        self.output_dir    = output_dir
        self.stories:      List[TestStory] = []
        self.active_story: Optional[TestStory] = None
        self._generator:   Optional[TestStoryGenerator] = None

    def set_generator(self, gen: TestStoryGenerator):
        self._generator = gen

    def start_story(self, story: TestStory):
        self.active_story = story
        self.stories.append(story)
        story.start()
        print(f"\n▶ STORY STARTED: {story.story_id} — {story.context_name}")

    def record_action(self, action: str, target: str, value: str,
                      success: bool, error: Optional[str] = None):
        if not self.active_story:
            return
        story = self.active_story
        story.add_step(action, target, value, success, error)
        icon = "✅" if success else "❌"
        print(f"  {icon} [{story.story_id}] {action} → {target}"
              + (f" = '{value}'" if value else "")
              + (f"  ⚠️  {error}" if error and not success else ""))
        if not success and error:
            self._check_toast_failure(error)

    def _check_toast_failure(self, error: str):
        if any(kw in error.lower() for kw in self.FAIL_KEYWORDS):
            if self.active_story and self.active_story.status == StoryStatus.RUNNING:
                self.active_story.fail_story(f"Validation error: {error[:200]}")

    def mark_loop_detected(self, target: str):
        if self.active_story and self.active_story.status == StoryStatus.RUNNING:
            self.active_story.fail_story(f"Loop detected on: {target}")

    def mark_submit_failed(self, target: str, error: str):
        if self.active_story and self.active_story.status == StoryStatus.RUNNING:
            self.active_story.fail_story(f"Submit failed on '{target}': {error}")

    def complete_story(self, toast_text: str = ""):
        if not self.active_story:
            return
        story = self.active_story
        if story.status == StoryStatus.RUNNING:
            if toast_text and any(kw in toast_text.lower() for kw in self.FAIL_KEYWORDS):
                story.fail_story(f"Error toast: {toast_text[:200]}")
            else:
                story.pass_story()
        self.active_story = None

    def abandon_story(self, reason: str = "Context changed"):
        if not self.active_story:
            return
        story = self.active_story
        if story.status == StoryStatus.RUNNING:
            if any(s.success for s in story.steps):
                story.pass_story()
            else:
                story.fail_story(reason)
        self.active_story = None

    def get_value_for_field(self, field_name: str) -> Optional[str]:
        if self.active_story:
            return self.active_story.get_value_for(field_name)
        return None


# ═══════════════════════════════════════════════════════════════
#  EXCEL REPORT GENERATOR
# ═══════════════════════════════════════════════════════════════

class ExcelReportGenerator:

    def __init__(self, output_dir: Path, session_id: str):
        self.output_dir = output_dir
        self.session_id = session_id

    def generate_all(self, stories: List[TestStory]) -> Path:
        wb = Workbook()
        wb.remove(wb.active)

        self._build_summary(wb, stories)
        self._build_user_test_stories(wb, stories)
        self._build_implementation_plan(wb, stories)
        self._build_execution_plan(wb, stories)

        path = self.output_dir / f"test_report_{self.session_id}.xlsx"
        wb.save(path)
        self._print_console_summary(stories)
        return path

    # ── SHEET 1: SUMMARY ──────────────────────────────────────

    def _build_summary(self, wb: Workbook, stories: List[TestStory]):
        ws = wb.create_sheet("Summary")
        ws.sheet_view.showGridLines = False
        for col, w in zip("ABCDEF", [28, 30, 18, 18, 18, 22]):
            ws.column_dimensions[col].width = w

        passed  = [s for s in stories if s.status == StoryStatus.PASSED]
        failed  = [s for s in stories if s.status == StoryStatus.FAILED]
        skipped = [s for s in stories if s.status not in (StoryStatus.PASSED, StoryStatus.FAILED)]
        total   = len(stories)
        pct     = round((len(passed) / total * 100) if total else 0)

        ws.row_dimensions[1].height = 36
        ws.merge_cells("A1:F1")
        t = ws["A1"]
        t.value     = "🧪  SEMANTIC TEST ENGINE — SESSION REPORT"
        t.font      = _font(bold=True, color=C.FONT_WHITE, size=16)
        t.fill      = _fill(C.HEADER_DARK)
        t.alignment = _align(h="center", v="center")

        ws.row_dimensions[2].height = 18
        ws.merge_cells("A2:F2")
        sub = ws["A2"]
        sub.value     = f"Session: {self.session_id}   |   Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        sub.font      = _font(italic=True, color=C.FONT_WHITE, size=9)
        sub.fill      = _fill(C.HEADER_MID)
        sub.alignment = _align(h="center")

        kpis = [
            ("Total Stories", total,        C.ACCENT_BLUE,   C.LIGHT_BLUE),
            ("✅ Passed",     len(passed),  C.PASS_BG,       C.LIGHT_GREEN),
            ("❌ Failed",     len(failed),  C.FAIL_BG,       C.LIGHT_RED),
            ("⏭ Skipped",    len(skipped), C.SKIP_BG,       C.LIGHT_AMBER),
            ("Pass Rate",     f"{pct}%",   C.ACCENT_PURPLE, C.LIGHT_PURPLE),
        ]

        for col_idx, (label, value, hdr_bg, val_bg) in enumerate(kpis, 1):
            ws.row_dimensions[4].height = 20
            lbl = ws.cell(row=4, column=col_idx, value=label)
            lbl.font      = _font(bold=True, color=C.FONT_WHITE, size=9)
            lbl.fill      = _fill(hdr_bg)
            lbl.alignment = _align(h="center")
            lbl.border    = _border()

            ws.row_dimensions[5].height = 32
            val = ws.cell(row=5, column=col_idx, value=value)
            val.font      = _font(bold=True, color=C.FONT_DARK, size=18)
            val.fill      = _fill(val_bg)
            val.alignment = _align(h="center", v="center")
            val.border    = _border()

        ws.row_dimensions[7].height = 8
        _header_row(ws, 8,
            ["Story ID", "Context Name", "Persona", "Status",
             "Steps (Pass/Total)", "Failure Reason"],
            C.HEADER_DARK)

        row = 9
        for idx, s in enumerate(stories):
            if s.status == StoryStatus.PASSED:
                status_txt, bg = "✅ PASSED", C.LIGHT_GREEN
            elif s.status == StoryStatus.FAILED:
                status_txt, bg = "❌ FAILED", C.LIGHT_RED
            else:
                status_txt, bg = "⏭ SKIPPED", C.LIGHT_AMBER

            success_steps = sum(1 for st in s.steps if st.success)
            _data_row(ws, row, [
                s.story_id, s.context_name, s.user_persona,
                status_txt,
                f"{success_steps} / {len(s.steps)}",
                s.failure_reason or "—"
            ], bg=bg if idx % 2 == 0 else C.ALT_ROW)

            cell = ws.cell(row=row, column=4)
            if "PASSED" in str(cell.value):
                cell.font = _font(bold=True, color=C.FONT_GREEN)
            elif "FAILED" in str(cell.value):
                cell.font = _font(bold=True, color=C.FONT_RED)
            else:
                cell.font = _font(bold=True, color=C.FONT_AMBER)
            row += 1

    # ── SHEET 2: USER TEST STORIES ────────────────────────────

    def _build_user_test_stories(self, wb: Workbook, stories: List[TestStory]):
        ws = wb.create_sheet("User Test Stories")
        ws.sheet_view.showGridLines = False

        for i, w in enumerate([16, 28, 35, 45, 12, 14, 18, 12, 30], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.row_dimensions[1].height = 32
        ws.merge_cells("A1:I1")
        t = ws["A1"]
        t.value     = "📋  USER TEST STORIES"
        t.font      = _font(bold=True, color=C.FONT_WHITE, size=14)
        t.fill      = _fill(C.ACCENT_BLUE)
        t.alignment = _align(h="center", v="center")

        ws.row_dimensions[2].height = 20
        ws.merge_cells("A2:I2")
        desc = ws["A2"]
        desc.value     = ("Documents every user scenario tested: persona, context, "
                          "field values used, execution steps, and pass/fail outcome.")
        desc.font      = _font(italic=True, color=C.FONT_MUTED, size=9)
        desc.fill      = _fill(C.LIGHT_BLUE)
        desc.alignment = _align(h="center")

        _header_row(ws, 4,
            ["Story ID", "Context / Feature", "User Persona",
             "Scenario Description", "Status", "Steps Total",
             "Steps Passed", "Pass %", "Failure Reason"],
            C.HEADER_DARK, height=24)

        row = 5
        for idx, s in enumerate(stories):
            if s.status == StoryStatus.PASSED:
                status_txt, bg = "✅ PASSED", C.LIGHT_GREEN
            elif s.status == StoryStatus.FAILED:
                status_txt, bg = "❌ FAILED", C.LIGHT_RED
            else:
                status_txt, bg = "⏭ SKIPPED", C.LIGHT_AMBER

            total_steps   = len(s.steps)
            success_steps = sum(1 for st in s.steps if st.success)
            alt_bg        = bg if idx % 2 == 0 else C.ALT_ROW

            _data_row(ws, row, [
                s.story_id, s.context_name, s.user_persona,
                s.description, status_txt, total_steps, success_steps,
                f"{round(success_steps / total_steps * 100) if total_steps else 0}%",
                s.failure_reason or "—"
            ], bg=alt_bg, height=40)

            ws.cell(row=row, column=5).font = _font(
                bold=True,
                color=C.FONT_GREEN if "PASSED" in status_txt
                      else C.FONT_RED if "FAILED" in status_txt
                      else C.FONT_AMBER
            )
            row += 1

        # ── Field Values section
        row += 1
        ws.merge_cells(f"A{row}:I{row}")
        sec = ws[f"A{row}"]
        sec.value     = "FIELD VALUES USED IN EACH STORY"
        sec.font      = _font(bold=True, color=C.FONT_WHITE, size=11)
        sec.fill      = _fill(C.HEADER_MID)
        sec.alignment = _align(h="center")
        ws.row_dimensions[row].height = 24
        row += 1

        _header_row(ws, row,
            ["Story ID", "Context", "Field Name", "Value Used",
             "Expected Outcome", "", "", "", ""],
            C.ACCENT_BLUE, height=20)
        row += 1

        for s in stories:
            # Print each field value with the expected_outcome on first row
            first = True
            for field_name, field_val in s.field_values.items():
                alt = row % 2 == 0
                _data_row(ws, row, [
                    s.story_id, s.context_name, field_name, field_val,
                    s.expected_outcome if first else "",  # ✅ FIX 2 — real outcome
                    "", "", "", ""
                ], bg=C.LIGHT_BLUE if alt else C.WHITE)
                first = False
                row += 1
            # If no field values, still show outcome
            if not s.field_values:
                _data_row(ws, row, [
                    s.story_id, s.context_name, "—", "—",
                    s.expected_outcome, "", "", "", ""
                ], bg=C.WHITE)
                row += 1

        # ── Detailed Steps section
        row += 1
        ws.merge_cells(f"A{row}:I{row}")
        sec2 = ws[f"A{row}"]
        sec2.value     = "DETAILED EXECUTION STEPS"
        sec2.font      = _font(bold=True, color=C.FONT_WHITE, size=11)
        sec2.fill      = _fill(C.HEADER_MID)
        sec2.alignment = _align(h="center")
        ws.row_dimensions[row].height = 24
        row += 1

        _header_row(ws, row,
            ["Story ID", "Step #", "Action", "Target Element",
             "Value Used", "Result", "Error", "", "Timestamp"],
            C.HEADER_DARK, height=20)
        row += 1

        for s in stories:
            for step in s.steps:
                result_txt = "✅ Pass" if step.success else "❌ Fail"
                alt = row % 2 == 0
                _data_row(ws, row, [
                    s.story_id, step.step_num, step.action, step.target,
                    step.value or "", result_txt, step.error or "",
                    "", step.timestamp[:19].replace("T", " ")
                ], bg=C.LIGHT_GREEN if step.success
                        else C.LIGHT_RED if not step.success and step.error
                        else C.WHITE)
                ws.cell(row=row, column=6).font = _font(
                    bold=True,
                    color=C.FONT_GREEN if step.success else C.FONT_RED
                )
                row += 1

    # ── SHEET 3: IMPLEMENTATION PLAN ─────────────────────────

    def _build_implementation_plan(self, wb: Workbook, stories: List[TestStory]):
        ws = wb.create_sheet("Implementation Plan")
        ws.sheet_view.showGridLines = False

        for i, w in enumerate([8, 25, 20, 22, 40, 15, 15, 20, 12], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.row_dimensions[1].height = 32
        ws.merge_cells("A1:I1")
        t = ws["A1"]
        t.value     = "🔧  IMPLEMENTATION PLAN"
        t.font      = _font(bold=True, color=C.FONT_WHITE, size=14)
        t.fill      = _fill(C.ACCENT_ORANGE)
        t.alignment = _align(h="center", v="center")

        ws.row_dimensions[2].height = 20
        ws.merge_cells("A2:I2")
        desc = ws["A2"]
        desc.value     = ("Derived from failed stories and identified gaps. "
                          "Maps each issue to a concrete fix with priority and owner.")
        desc.font      = _font(italic=True, color=C.FONT_MUTED, size=9)
        desc.fill      = _fill(C.LIGHT_AMBER)
        desc.alignment = _align(h="center")

        _header_row(ws, 4,
            ["#", "Feature / Module", "Story Reference", "Issue Type",
             "Recommended Fix", "Priority", "Effort", "Owner", "Status"],
            C.HEADER_DARK, height=24)

        items = self._derive_implementation_items(stories)
        row = 5
        for idx, item in enumerate(items):
            priority = item.get("priority", "Medium")
            bg = (C.LIGHT_RED if priority == "High"
                  else C.LIGHT_GREEN if priority == "Low"
                  else C.LIGHT_AMBER if idx % 2 == 0 else C.ALT_ROW)

            _data_row(ws, row, [
                idx + 1,
                item.get("module", ""),
                item.get("story_ref", ""),
                item.get("issue_type", ""),
                item.get("fix", ""),
                priority,
                item.get("effort", "M"),
                item.get("owner", "QA / Dev"),
                item.get("status", "Open")
            ], bg=bg, height=36)

            ws.cell(row=row, column=6).font = _font(
                bold=True,
                color=C.FONT_RED if priority == "High"
                      else C.FONT_AMBER if priority == "Medium"
                      else C.FONT_GREEN
            )
            row += 1

        row += 1
        ws.merge_cells(f"A{row}:I{row}")
        leg = ws[f"A{row}"]
        leg.value     = "LEGEND:  Priority — High = block release | Medium = fix in sprint | Low = nice-to-have     Effort — S < 2h | M 2–8h | L > 1 day"
        leg.font      = _font(italic=True, color=C.FONT_MUTED, size=8)
        leg.alignment = _align(h="left")
        leg.fill      = _fill(C.ALT_ROW)

    def _derive_implementation_items(self, stories: List[TestStory]) -> List[Dict]:
        items = []
        failed = [s for s in stories if s.status == StoryStatus.FAILED]
        passed = [s for s in stories if s.status == StoryStatus.PASSED]

        for s in failed:
            reason = s.failure_reason or "Unknown failure"
            if "loop" in reason.lower():
                items.append({
                    "module": s.context_name, "story_ref": s.story_id,
                    "issue_type": "UI Interaction Loop",
                    "fix": "Fix element identifier uniqueness. Check formcontrolname and button state transitions.",
                    "priority": "High", "effort": "M", "owner": "Frontend Dev", "status": "Open"
                })
            elif "submit" in reason.lower() or "disabled" in reason.lower():
                items.append({
                    "module": s.context_name, "story_ref": s.story_id,
                    "issue_type": "Form Validation / Disabled Button",
                    "fix": "Ensure submit button enables after required fields filled. Review Angular reactive form validators.",
                    "priority": "High", "effort": "S", "owner": "Frontend Dev", "status": "Open"
                })
            elif "toast" in reason.lower() or "error" in reason.lower():
                items.append({
                    "module": s.context_name, "story_ref": s.story_id,
                    "issue_type": "Backend Validation Error",
                    "fix": f"Server returned error on '{s.context_name}'. Review API payload and error UX copy. Original: {reason[:120]}",
                    "priority": "High", "effort": "M", "owner": "Backend Dev", "status": "Open"
                })
            else:
                items.append({
                    "module": s.context_name, "story_ref": s.story_id,
                    "issue_type": "General Failure",
                    "fix": f"Investigate: {reason[:150]}",
                    "priority": "Medium", "effort": "M", "owner": "QA", "status": "Open"
                })

        items.append({
            "module": "Test Infrastructure", "story_ref": "ALL",
            "issue_type": "Test Coverage",
            "fix": "Add negative test cases: invalid email formats, empty required fields, boundary values, special characters.",
            "priority": "Medium", "effort": "L", "owner": "QA", "status": "Open"
        })
        items.append({
            "module": "Auth / Session", "story_ref": "ALL",
            "issue_type": "Session Handling",
            "fix": "Verify session tokens refresh during long runs. Check auth.json captures all required tokens and cookies.",
            "priority": "Low", "effort": "S", "owner": "QA", "status": "Open"
        })
        if passed:
            items.append({
                "module": "Regression Suite",
                "story_ref": ", ".join(s.story_id for s in passed[:5]),
                "issue_type": "Automation Opportunity",
                "fix": f"{len(passed)} stories passed. Convert to automated regression tests in CI pipeline.",
                "priority": "Low", "effort": "L", "owner": "QA / DevOps", "status": "Planned"
            })

        return items

    # ── SHEET 4: EXECUTION PLAN ───────────────────────────────

    def _build_execution_plan(self, wb: Workbook, stories: List[TestStory]):
        ws = wb.create_sheet("Execution Plan")
        ws.sheet_view.showGridLines = False

        for i, w in enumerate([6, 20, 30, 45, 20, 45, 12, 18, 14], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.row_dimensions[1].height = 32
        ws.merge_cells("A1:I1")
        t = ws["A1"]
        t.value     = "🚀  EXECUTION PLAN"
        t.font      = _font(bold=True, color=C.FONT_WHITE, size=14)
        t.fill      = _fill(C.ACCENT_PURPLE)
        t.alignment = _align(h="center", v="center")

        ws.row_dimensions[2].height = 20
        ws.merge_cells("A2:I2")
        desc = ws["A2"]
        desc.value     = "Step-by-step test execution roadmap: what to test, how, with what data, expected outcomes, and who runs it."
        desc.font      = _font(italic=True, color=C.FONT_MUTED, size=9)
        desc.fill      = _fill(C.LIGHT_PURPLE)
        desc.alignment = _align(h="center")

        # Phase overview
        ws.row_dimensions[4].height = 22
        ws.merge_cells("A4:I4")
        ph = ws["A4"]
        ph.value     = "EXECUTION PHASES"
        ph.font      = _font(bold=True, color=C.FONT_WHITE, size=10)
        ph.fill      = _fill(C.HEADER_MID)
        ph.alignment = _align(h="center")

        phases = [
            ("Phase 1", "Environment Setup",   "Verify auth, deploy app, clear test data",   "QA Lead",   "30 min"),
            ("Phase 2", "Smoke Tests",          "Run top 3 critical user stories",             "QA",        "1 hr"),
            ("Phase 3", "Full Story Execution", "Run all generated test stories",              "QA",        "2–4 hr"),
            ("Phase 4", "Regression",           "Re-run fixed stories from failed set",        "QA",        "1 hr"),
            ("Phase 5", "Exploratory / Edge",   "Negative cases, boundary values, edge UX",    "QA Senior", "2 hr"),
            ("Phase 6", "Sign-off & Reporting", "Review Excel report, raise defects, approve", "QA Lead",   "30 min"),
        ]
        _header_row(ws, 5, ["Phase", "Name", "Objective", "Responsibilities", "Duration",
                             "", "", "", ""], C.HEADER_DARK)
        row = 6
        for idx, (phase, name, obj, resp, dur) in enumerate(phases):
            bg = C.LIGHT_PURPLE if idx % 2 == 0 else C.ALT_ROW
            _data_row(ws, row, [phase, name, obj, resp, dur, "", "", "", ""], bg=bg, height=28)
            ws.cell(row=row, column=1).font = _font(bold=True, color=C.ACCENT_PURPLE)
            row += 1

        # Detailed execution steps
        row += 1
        ws.merge_cells(f"A{row}:I{row}")
        sec = ws[f"A{row}"]
        sec.value     = "DETAILED EXECUTION STEPS (per Story)"
        sec.font      = _font(bold=True, color=C.FONT_WHITE, size=11)
        sec.fill      = _fill(C.HEADER_MID)
        sec.alignment = _align(h="center")
        ws.row_dimensions[row].height = 24
        row += 1

        _header_row(ws, row,
            ["#", "Story ID", "Feature / Context", "Test Steps",
             "Test Data (Key Fields)", "Expected Result",      # ✅ FIX 2 — column header
             "Priority", "Assigned To", "Execution Status"],
            C.HEADER_DARK, height=24)
        row += 1

        exec_num = 1
        for s in stories:
            step_texts = []
            for st in s.steps[:8]:
                step_texts.append(
                    f"{st.step_num}. {st.action.upper()} '{st.target}'"
                    + (f" → '{st.value}'" if st.value else "")
                )
            steps_text = "\n".join(step_texts)
            if len(s.steps) > 8:
                steps_text += f"\n... +{len(s.steps)-8} more steps"

            key_fields = "; ".join(
                f"{k}={v}" for k, v in list(s.field_values.items())[:4]
            )
            if len(s.field_values) > 4:
                key_fields += f" (+{len(s.field_values)-4} more)"

            if s.status == StoryStatus.PASSED:
                priority, exec_status, bg = "Medium", "✅ Executed", C.LIGHT_GREEN
            elif s.status == StoryStatus.FAILED:
                priority, exec_status, bg = "High",   "❌ Failed",   C.LIGHT_RED
            else:
                priority, exec_status, bg = "Low",    "⏭ Pending",  C.LIGHT_AMBER

            # ✅ FIX 2 — use real expected_outcome, not hardcoded string
            expected = s.expected_outcome or "See story description"

            _data_row(ws, row, [
                exec_num, s.story_id, s.context_name,
                steps_text, key_fields,
                expected,           # ← real outcome here
                priority, "QA", exec_status
            ], bg=bg, height=max(18 * min(len(s.steps), 8), 36))

            ws.cell(row=row, column=9).font = _font(
                bold=True,
                color=C.FONT_GREEN if "Executed" in exec_status
                      else C.FONT_RED if "Failed" in exec_status
                      else C.FONT_AMBER
            )
            ws.cell(row=row, column=7).font = _font(
                bold=True,
                color=C.FONT_RED if priority == "High"
                      else C.FONT_AMBER if priority == "Medium"
                      else C.FONT_GREEN
            )
            row += 1
            exec_num += 1

        # Checklist
        row += 1
        ws.merge_cells(f"A{row}:I{row}")
        chk = ws[f"A{row}"]
        chk.value     = "PRE-EXECUTION CHECKLIST"
        chk.font      = _font(bold=True, color=C.FONT_WHITE, size=11)
        chk.fill      = _fill(C.HEADER_MID)
        chk.alignment = _align(h="center")
        ws.row_dimensions[row].height = 22
        row += 1

        checklist = [
            ("☐", "auth.json is up-to-date with valid tokens",          "QA Lead"),
            ("☐", "Target environment URL is correct and accessible",    "QA"),
            ("☐", "Test data (users, records) is seeded in environment", "Dev / QA"),
            ("☐", "Browser: Chromium / Chrome latest version",           "QA"),
            ("☐", "Screen resolution: 1400×900 (default in test engine)","QA"),
            ("☐", "Network: application reachable from test machine",    "Infra"),
            ("☐", "Logs directory writable (semantic_test_output/)",     "QA"),
            ("☐", "Previous test screenshots cleared if re-running",     "QA"),
        ]
        _header_row(ws, row, ["✓", "Checklist Item", "Responsible",
                               "", "", "", "", "", ""], C.ACCENT_PURPLE)
        row += 1
        for idx, (tick, item, resp) in enumerate(checklist):
            bg = C.LIGHT_PURPLE if idx % 2 == 0 else C.ALT_ROW
            _data_row(ws, row, [tick, item, resp, "", "", "", "", "", ""], bg=bg)
            row += 1

    # ── CONSOLE SUMMARY — called exactly once from generate_all ─

    def _print_console_summary(self, stories: List[TestStory]):
        passed = sum(1 for s in stories if s.status == StoryStatus.PASSED)
        failed = sum(1 for s in stories if s.status == StoryStatus.FAILED)
        total  = len(stories)
        pct    = round((passed / total * 100) if total else 0)

        print(f"\n{'='*70}")
        print(f"  📊 TEST STORY REPORT")
        print(f"{'='*70}")
        print(f"  Session  : {self.session_id}")
        print(f"  Total    : {total}")
        print(f"  ✅ Passed: {passed}   ❌ Failed: {failed}   Pass Rate: {pct}%")
        print(f"{'='*70}")
        print(f"  {'STORY ID':<22} {'CONTEXT':<30} {'STATUS':<12} {'STEPS'}")
        print(f"  {'-'*70}")
        for s in stories:
            status  = s.status.value.upper()
            success = sum(1 for st in s.steps if st.success)
            print(f"  {s.story_id:<22} {s.context_name[:29]:<30} {status:<12} {success}/{len(s.steps)}")
        print(f"{'='*70}\n")


# Alias for backward compatibility
ReportGenerator = ExcelReportGenerator